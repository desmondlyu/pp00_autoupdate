import re
import tkinter as tk
from tkinter import filedialog, font
from openpyxl import load_workbook
import os
from pathlib import Path
import threading
import time
from tkinter import ttk

# 原始 key(lower) -> mapped title
MAPPING = {
    'spec':           'DC_spec(uA)',
    'speci':          'DC_spec(uA)',
    'gate_val':       'Force(CSb)',
    'well_val':       'Force(CSb)',
    'cs_val':         'Force(CSb)',
    'vg':             'Force(CSb)',
    'drain_val':      'Force(SO)',
    'x1_addr':        'Start_Address',
    'wl_add1':        'Start_Address',
    'ifr_uid':        'Start_Address',
    'x2_addr':        'End_Address',
    'ifr_uid_addr':   'End_Address',
    'wl_add2':        'End_Address',
    'opt31_0':        'Data_1',
    'data':           'Data_1',
    'ldata':          'Data_1',
    'data1':          'Data_1',
    'ifr_uid_data':   'Data_1',
    'opt63_32':       'Data_2',
    'hdata':          'Data_2',
    'data2':          'Data_2',
    'opt95_64':       'Data_3',
    'tout':           'Time_out',
    'erase_step':     'Time_out',
    'twp':            'tWC',
    'rc':             'tWC',
    'twp_val':        'tWC',
    'pulse':          'Pulse',
    'flag':           'MR_Ratio',
    'flag1':          'MR_Ratio',
    'flag2':          'MR_Ratio',
    'mr_flag':        'MR_Ratio',
}

# 僅保留這 12 個標題，且以此順序顯示
MAPPED_TITLES = [
    'DC_spec(uA)', 'Force(CSb)', 'Force(SO)', 'Start_Address', 'End_Address',
    'Data_1', 'Data_2', 'Data_3', 'MR_Ratio', 'Time_out',  'tWC', 'Pulse'
]

# MR 欄位只接受 MR11(0) 或 MR12(1)，不分大小寫，可有可無空格
MR_PATTERN = re.compile(r'^(mr11\s*\(0\)|mr12\s*\(1\))$', re.IGNORECASE)
# BE_TIME 或 SE_TIME 的擷取模式
TIMEOUT_PATTERN = re.compile(r'^(?:BE_TIME|SE_TIME)\s*\(\s*(\d+)\s*\)$', re.IGNORECASE)

class TooltipBase:
    def __init__(self, widget, text, delay=500, **kwargs):
        self.widget = widget
        self.text = text
        self.delay = delay
        
        self.tooltip = None
        self.id = None
        
        self.style_kwargs = {
            'bg': '#2a2a2a',
            'fg': '#e0e0e0',
            'padx': 10,
            'pady': 5,
            'bd': 0,
            'relief': 'solid'
        }
        self.style_kwargs.update(kwargs)
        
        self.widget.bind("<Enter>", self.schedule)
        self.widget.bind("<Leave>", self.hide)
        self.widget.bind("<Button-1>", self.hide)
    
    def schedule(self, event=None):
        self.id = self.widget.after(self.delay, self.show)
    
    def show(self, event=None):
        self.hide()
        
        x, y, _, _ = self.widget.bbox("insert")
        x += self.widget.winfo_rootx() + 25
        y += self.widget.winfo_rooty() + 25
        
        self.tooltip = tk.Toplevel(self.widget)
        self.tooltip.wm_overrideredirect(True)
        self.tooltip.wm_geometry(f"+{x}+{y}")
        
        label = tk.Label(self.tooltip, text=self.text, justify='left', **self.style_kwargs)
        label.pack()
    
    def hide(self, event=None):
        if self.id:
            self.widget.after_cancel(self.id)
            self.id = None
        
        if self.tooltip:
            self.tooltip.destroy()
            self.tooltip = None

class MacOSButton(tk.Canvas):
    def __init__(self, master, text="", command=None, width=200, height=34, corner_radius=6, 
                 bg='#333333', fg='#ffffff', hover_color='#404040', **kwargs):
        super().__init__(master, width=width, height=height, bg=bg, bd=0, 
                         highlightthickness=0, relief="ridge", **kwargs)
        
        self.command = command
        self.corner_radius = corner_radius
        self.bg = bg
        self.fg = fg
        self.hover_color = hover_color
        self.configure(cursor="hand2")
        
        # Initial button state
        self.button_state = "normal"
        
        # Draw the rounded rectangle button
        self.rect = self.create_rounded_rect(0, 0, width, height, corner_radius, fill=bg, outline="")
        self.text_id = self.create_text(width/2, height/2, text=text, fill=fg, font=("SF Pro Text", 12))
        
        # Bind events
        self.bind("<Enter>", self.on_enter)
        self.bind("<Leave>", self.on_leave)
        self.bind("<ButtonPress-1>", self.on_press)
        self.bind("<ButtonRelease-1>", self.on_release)

    def create_rounded_rect(self, x1, y1, x2, y2, radius, **kwargs):
        points = [
            x1+radius, y1,
            x1+radius, y1,
            x2-radius, y1,
            x2-radius, y1,
            x2, y1,
            x2, y1+radius,
            x2, y1+radius,
            x2, y2-radius,
            x2, y2-radius,
            x2, y2,
            x2-radius, y2,
            x2-radius, y2,
            x1+radius, y2,
            x1+radius, y2,
            x1, y2,
            x1, y2-radius,
            x1, y2-radius,
            x1, y1+radius,
            x1, y1+radius,
            x1, y1
        ]
        return self.create_polygon(points, **kwargs, smooth=True)

    def on_enter(self, e):
        self.itemconfig(self.rect, fill=self.hover_color)

    def on_leave(self, e):
        self.itemconfig(self.rect, fill=self.bg)

    def on_press(self, e):
        self.itemconfig(self.rect, fill="#555555")  # Darker when pressed

    def on_release(self, e):
        self.itemconfig(self.rect, fill=self.hover_color)
        if self.command:
            self.command()

class ProgressDialog(tk.Toplevel):
    def __init__(self, parent, title="處理中"):
        super().__init__(parent)
        self.title(title)
        self.configure(bg="#2a2a2a")
        self.resizable(False, False)
        self.transient(parent)
        self.grab_set()
        
        # Calculate position to center on parent
        parent_x = parent.winfo_rootx()
        parent_y = parent.winfo_rooty()
        parent_width = parent.winfo_width()
        parent_height = parent.winfo_height()
        
        width = 300
        height = 80
        x = parent_x + (parent_width - width) // 2
        y = parent_y + (parent_height - height) // 2
        
        self.geometry(f"{width}x{height}+{x}+{y}")
        
        self.label = tk.Label(self, text="正在處理檔案...", bg="#2a2a2a", fg="#ffffff", font=("SF Pro Text", 12))
        self.label.pack(pady=(10, 0))
        
        self.progress = ttk.Progressbar(self, orient="horizontal", mode="indeterminate", length=250)
        self.progress.pack(pady=15, padx=25)
        self.progress.start(10)
        
    def update_status(self, text):
        self.label.config(text=text)
        self.update()

class MacOSAlert(tk.Toplevel):
    def __init__(self, parent, title, message, icon_type="info"):
        super().__init__(parent)
        self.title("")
        self.configure(bg="#2a2a2a")
        self.resizable(False, False)
        self.transient(parent)
        
        # No window decorations on macOS-style dialogs
        self.overrideredirect(True)
        
        # Calculate position to center on parent
        parent_x = parent.winfo_rootx()
        parent_y = parent.winfo_rooty()
        parent_width = parent.winfo_width()
        parent_height = parent.winfo_height()
        
        width = 400
        height = 170
        x = parent_x + (parent_width - width) // 2
        y = parent_y + (parent_height - height) // 2
        
        self.geometry(f"{width}x{height}+{x}+{y}")
        
        # Add drop shadow effect (simulated with a frame)
        shadow_frame = tk.Frame(self, bg="#1a1a1a", bd=0)
        shadow_frame.place(x=3, y=3, width=width, height=height)
        
        # Main content frame
        main_frame = tk.Frame(self, bg="#2a2a2a", bd=0)
        main_frame.place(x=0, y=0, width=width, height=height)
        
        # Icon and title frame
        header_frame = tk.Frame(main_frame, bg="#2a2a2a", height=40)
        header_frame.pack(fill=tk.X, pady=(15, 5))
        
        # Icon (represented as text for simplicity)
        if icon_type == "info":
            icon_text = "ℹ️"
        elif icon_type == "warning":
            icon_text = "⚠️"
        elif icon_type == "error":
            icon_text = "❌"
        else:
            icon_text = "ℹ️"
            
        icon_label = tk.Label(header_frame, text=icon_text, bg="#2a2a2a", fg="#ffffff", font=("SF Pro Text", 24))
        icon_label.pack(side=tk.LEFT, padx=(20, 0))
        
        # Title
        title_label = tk.Label(header_frame, text=title, bg="#2a2a2a", fg="#ffffff", font=("SF Pro Text", 16, "bold"))
        title_label.pack(side=tk.LEFT, padx=10)
        
        # Message
        message_frame = tk.Frame(main_frame, bg="#2a2a2a")
        message_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=5)
        
        message_label = tk.Label(message_frame, text=message, bg="#2a2a2a", fg="#e0e0e0", 
                                font=("SF Pro Text", 12), wraplength=360, justify="left")
        message_label.pack(fill=tk.BOTH, expand=True)
        
        # Button frame
        button_frame = tk.Frame(main_frame, bg="#2a2a2a", height=50)
        button_frame.pack(fill=tk.X, pady=(0, 15))
        
        self.ok_button = MacOSButton(button_frame, text="確定", command=self.destroy, 
                                    width=80, height=30, bg="#0066cc", hover_color="#0077ee")
        self.ok_button.pack(side=tk.RIGHT, padx=(0, 20))
        
        # Make dialog modal
        self.grab_set()
        self.focus_set()
        
        # Position the window in the center of the parent
        self.update_idletasks()
        
        # Add a subtle bounce animation
        self.animate_entrance()
        
        # Bind Escape key to close
        self.bind("<Escape>", lambda e: self.destroy())
        
    def animate_entrance(self):
        # Store original position
        orig_y = self.winfo_y()
        
        # Start slightly above
        self.geometry(f"+{self.winfo_x()}+{orig_y-15}")
        self.update_idletasks()
        
        # Bounce down
        def bounce_down():
            for i in range(15):
                self.geometry(f"+{self.winfo_x()}+{self.winfo_y()+1}")
                self.update_idletasks()
                time.sleep(0.01)
        
        threading.Thread(target=bounce_down, daemon=True).start()

def extract_comments_all_sheets(
    file_path: str,
    comment_col: int = 7,    # G 欄 (1=A,2=B…,7=G) 為註解來源
    start_row: int = 3,      # 從 G3 開始掃描
    header_row: int = 1      # 標題列在第 1 列
):
    wb = load_workbook(filename=file_path)
    entry_pattern = re.compile(r'^(\w+)\s*=\s*(.+)$')

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        print(f"處理工作表：{sheet}")

        # 刪除 H 欄之後所有欄
        if ws.max_column >= 8:
            ws.delete_cols(8, ws.max_column - 7)

        # 刪除 A 欄空白行
        for r in range(ws.max_row, header_row, -1):
            if not ws.cell(row=r, column=1).value:
                ws.delete_rows(r)

        # 重置並寫入固定 12 個標題
        for col in range(comment_col+1, ws.max_column+1):
            ws.cell(row=header_row, column=col, value=None)
        header_map = {}
        for idx, title in enumerate(MAPPED_TITLES):
            col = comment_col + 1 + idx
            ws.cell(row=header_row, column=col, value=title)
            header_map[title.lower()] = col

        # 解析每列註解
        row = start_row
        while True:
            if not ws.cell(row=row, column=1).value:
                break

            comment = ws.cell(row=row, column=comment_col).comment
            if not comment or not comment.text.strip():
                row += 1
                continue

            entries = []
            for ln in comment.text.splitlines()[1:]:
                ln = ln.strip()
                m = entry_pattern.match(ln)
                if m:
                    entries.append((m.group(1).lower(), m.group(2).strip()))

            keys_low = {k for k, _ in entries}
            # RC override 條件
            override_rc = None
            if 'rc' in keys_low:
                if any(k.startswith('twp') for k in keys_low):
                    override_rc = 'Pulse'
                elif 'pulse' in keys_low:
                    override_rc = 'tWC'

            for key_low, val in entries:
                # 規則：BE_TIME / SE_TIME 轉為 tout
                if key_low == 'spec':
                    tm = TIMEOUT_PATTERN.match(val)
                    if tm:
                        mapped = 'Time_out'
                        val = tm.group(1)  # 擷取括號內數字
                    else:
                        mapped = MAPPING.get(key_low)
                # MR Ratio 專用：只 accept MR11(0) / MR12(1)
                elif key_low in {'mr_flag', 'flag', 'flag1', 'flag2'}:
                    if not MR_PATTERN.match(val):
                        continue
                    mapped = MAPPING.get(key_low)
                    val = val.upper().replace('MR11(', 'MR11 (').replace('MR12(', 'MR12 (')
                # RC override
                elif key_low == 'rc' and override_rc:
                    mapped = override_rc
                else:
                    mapped = MAPPING.get(key_low)

                if not mapped:
                    continue

                col = header_map[mapped.lower()]
                ws.cell(row=row, column=col, value=val)

            row += 1

    wb.save(file_path)
    print("所有工作表處理完成並已儲存。")

class MainApplication(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("MSS Transfer 工具")
        self.configure(bg="#1e1e1e")  # Dark background
        self.minsize(550, 350)
        self.geometry("550x350")
        
        # Set system font
        self.system_font = font.nametofont("TkDefaultFont")
        self.system_font.configure(family="SF Pro Text", size=12)
        
        # Setting ttk style
        self.style = ttk.Style()
        self.style.theme_use('default')
        self.style.configure("TProgressbar", thickness=6, background='#0066cc')
        
        self.create_widgets()
        self.center_window()
        self.current_file = None
        
    def create_widgets(self):
        # Main frame with padding
        main_frame = tk.Frame(self, bg="#1e1e1e", padx=20, pady=20)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Header Frame
        header_frame = tk.Frame(main_frame, bg="#1e1e1e")
        header_frame.pack(fill=tk.X, pady=(0, 15))
        
        # Title with SF Pro Display font
        title_label = tk.Label(header_frame, text="MSS 資料轉換器", 
                               font=("SF Pro Display", 18, "bold"), 
                               bg="#1e1e1e", fg="#ffffff")
        title_label.pack(side=tk.LEFT)
        
        # Information icon with tooltip
        info_frame = tk.Frame(header_frame, bg="#1e1e1e")
        info_frame.pack(side=tk.RIGHT, padx=5)
        
        info_label = tk.Label(info_frame, text="ⓘ", font=("SF Pro Text", 14), 
                             bg="#1e1e1e", fg="#4a90e2", cursor="hand2")
        info_label.pack()
        
        # Tooltip for info button
        tooltip_text = ("作者: PP32 YPLu + AI\n"
                        "版本: 20250529_00")
        TooltipBase(info_label, tooltip_text)
        
        # Description frame
        desc_frame = tk.Frame(main_frame, bg="#1e1e1e")
        desc_frame.pack(fill=tk.X, pady=(0, 20))
        
        desc_text = ("MSS 註解轉換 for PE\n"
                       "• 請確保你的G欄位是MSS註解\n"
                       "• 自動匯入固定12格PE必填欄位\n"
                       "• 支援多個測試站點分頁整理")
        desc_label = tk.Label(desc_frame, text=desc_text, bg="#1e1e1e", fg="#a0a0a0", 
                             font=("SF Pro Text", 11), justify=tk.LEFT)
        desc_label.pack(anchor="w")
        
        # File selection frame
        file_frame = tk.Frame(main_frame, bg="#1e1e1e")
        file_frame.pack(fill=tk.X, pady=10)
        
        # File path display with ellipsis for long paths
        self.file_var = tk.StringVar()
        self.file_var.set("尚未選擇檔案")
        
        file_label = tk.Label(file_frame, text="檔案路徑:", bg="#1e1e1e", fg="#e0e0e0", 
                             font=("SF Pro Text", 12))
        file_label.pack(side=tk.LEFT, padx=(0, 5))
        
        self.path_label = tk.Label(file_frame, textvariable=self.file_var, bg="#1e1e1e", 
                                  fg="#cccccc", anchor="w", width=40, font=("SF Pro Text", 12))
        self.path_label.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        # Button frame
        button_frame = tk.Frame(main_frame, bg="#1e1e1e")
        button_frame.pack(fill=tk.X, pady=20)
        
        # Open file button
        self.open_button = MacOSButton(button_frame, text="選擇 MSS 檔案", command=self.select_file,
                                      width=150, height=34, bg="#333333", hover_color="#404040")
        self.open_button.pack(side=tk.LEFT, padx=(0, 10))
        
        # Process button
        self.process_button = MacOSButton(button_frame, text="開始處理", command=self.process_file,
                                         width=150, height=34, bg="#0066cc", hover_color="#0077ee")
        self.process_button.pack(side=tk.LEFT)
        self.process_button.configure(state=tk.DISABLED)  # Initially disabled
        
        # Status frame at the bottom
        status_frame = tk.Frame(main_frame, bg="#252525", bd=0, height=30)
        status_frame.pack(fill=tk.X, side=tk.BOTTOM, pady=(20, 0))
        
        self.status_var = tk.StringVar()
        self.status_var.set("準備就緒")
        
        status_label = tk.Label(status_frame, textvariable=self.status_var, bg="#252525", 
                               fg="#a0a0a0", anchor="w", padx=10, pady=5, font=("SF Pro Text", 10))
        status_label.pack(fill=tk.X)

    def center_window(self):
        """Center the window on the screen"""
        self.update_idletasks()
        width = self.winfo_width()
        height = self.winfo_height()
        x = (self.winfo_screenwidth() // 2) - (width // 2)
        y = (self.winfo_screenheight() // 2) - (height // 2)
        self.geometry(f'{width}x{height}+{x}+{y}')
    
    def select_file(self):
        filepath = filedialog.askopenfilename(
            title="選擇 Excel 檔案",
            filetypes=[("Excel 活頁簿", "*.xlsx *.xlsm"), ("所有檔案", "*.*")],
            initialdir=os.path.expanduser("~/Documents")  # Default to Documents folder
        )
        
        if not filepath:
            return
            
        # Update file path display with ellipsis for long paths
        if len(filepath) > 40:
            display_path = filepath[:18] + "..." + filepath[-19:]
        else:
            display_path = filepath
            
        self.file_var.set(display_path)
        self.current_file = filepath
        
        # Enable process button
        self.process_button.configure(state=tk.NORMAL)
        self.status_var.set(f"已載入: {Path(filepath).name}")

    def process_file_thread(self, filepath, progress_dialog):
        try:
            # Update status
            progress_dialog.update_status("正在處理 MSS 檔案...")
            
            # Process the file using the original logic
            extract_comments_all_sheets(
                file_path=filepath,
                comment_col=7,   # G 欄
                start_row=3,     # 從第 3 列
                header_row=1
            )
            
            # Close progress dialog
            progress_dialog.destroy()
            
            # Show success message
            success_message = f"已成功處理檔案：\n{Path(filepath).name}\n\n所有工作表的註解資料已轉換完成。"
            MacOSAlert(self, "完成", success_message, "info")
            
            # Update status
            self.status_var.set("處理完成")
            
        except Exception as e:
            # Close progress dialog before showing error
            progress_dialog.destroy()
            MacOSAlert(self, "錯誤", f"處理檔案時發生錯誤：\n{str(e)}", "error")
            self.status_var.set("處理時發生錯誤")

    def process_file(self):
        if not self.current_file:
            MacOSAlert(self, "注意", "請先選擇一個 Excel 檔案。", "warning")
            return
            
        # Show progress dialog
        progress_dialog = ProgressDialog(self, "處理中")
        
        # Process in a separate thread to keep UI responsive
        threading.Thread(
            target=self.process_file_thread, 
            args=(self.current_file, progress_dialog),
            daemon=True
        ).start()

# Set macOS-style appearance for the application
def set_macos_appearance():
    try:
        # For macOS
        import platform
        if platform.system() == 'Darwin':
            os.system('''defaults write -g NSRequiresAquaSystemAppearance -bool YES''')
            from tkmacosx import ColorVar, ColorscaleVar
    except:
        pass

if __name__ == '__main__':
    set_macos_appearance()
    app = MainApplication()
    app.mainloop()