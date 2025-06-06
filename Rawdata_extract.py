import tkinter as tk
from tkinter import filedialog, font
import re
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import os
from pathlib import Path
import threading
import time
from tkinter import ttk

# Excel 表頭欄位
HEADERS = [
    "Item", "Detail", "Vcc", "Vhh", "ICC1", "ICC1_POR", "ICC2", "ICC3", "Vih", "pin", "CS", "I", "D", "D1", "D2", "RC",
    "Tspec", "T", "Gate", "Twp", "Drain", "X", "Y", "X1", "X2", "OPT[31:0]", "OPT[63:32]", "OPT[87:64]", "OPT[95:88]",
    "Tbusy", "FB", "PR", "ERS", "PST", "REF", "All", "Twc", "Terase",
    "UID1_d0", "UID1_d1", "UID1_d2", "UID1_d3", "UID2_d0", "UID2_d1", "UID2_d2", "UID2_d3",
]

keyword_to_header = {k.strip("=").strip(): k.strip("=").strip() for k in HEADERS[2:]}  # 忽略前兩欄 Item, Detail

class TooltipBase:
    def __init__(self, widget, text, delay=500, **kwargs):
        self.widget = widget
        self.text = text
        self.delay = delay
        
        self.tooltip = None
        self.id = None
        
        self.style_kwargs = {
            'bg': '#2a2a2a',
            'fg': '#e0e0e0',
            'padx': 10,
            'pady': 5,
            'bd': 0,
            'relief': 'solid'
        }
        self.style_kwargs.update(kwargs)
        
        self.widget.bind("<Enter>", self.schedule)
        self.widget.bind("<Leave>", self.hide)
        self.widget.bind("<Button-1>", self.hide)
    
    def schedule(self, event=None):
        self.id = self.widget.after(self.delay, self.show)
    
    def show(self, event=None):
        self.hide()
        
        x, y, _, _ = self.widget.bbox("insert")
        x += self.widget.winfo_rootx() + 25
        y += self.widget.winfo_rooty() + 25
        
        self.tooltip = tk.Toplevel(self.widget)
        self.tooltip.wm_overrideredirect(True)
        self.tooltip.wm_geometry(f"+{x}+{y}")
        
        label = tk.Label(self.tooltip, text=self.text, justify='left', **self.style_kwargs)
        label.pack()
    
    def hide(self, event=None):
        if self.id:
            self.widget.after_cancel(self.id)
            self.id = None
        
        if self.tooltip:
            self.tooltip.destroy()
            self.tooltip = None

class MacOSButton(tk.Canvas):
    def __init__(self, master, text="", command=None, width=200, height=34, corner_radius=6, 
                 bg='#333333', fg='#ffffff', hover_color='#404040', **kwargs):
        super().__init__(master, width=width, height=height, bg=bg, bd=0, 
                         highlightthickness=0, relief="ridge", **kwargs)
        
        self.command = command
        self.corner_radius = corner_radius
        self.bg = bg
        self.fg = fg
        self.hover_color = hover_color
        self.configure(cursor="hand2")
        
        # Initial button state
        self.button_state = "normal"
        
        # Draw the rounded rectangle button
        self.rect = self.create_rounded_rect(0, 0, width, height, corner_radius, fill=bg, outline="")
        self.text_id = self.create_text(width/2, height/2, text=text, fill=fg, font=("SF Pro Text", 12))
        
        # Bind events
        self.bind("<Enter>", self.on_enter)
        self.bind("<Leave>", self.on_leave)
        self.bind("<ButtonPress-1>", self.on_press)
        self.bind("<ButtonRelease-1>", self.on_release)

    def create_rounded_rect(self, x1, y1, x2, y2, radius, **kwargs):
        points = [
            x1+radius, y1,
            x1+radius, y1,
            x2-radius, y1,
            x2-radius, y1,
            x2, y1,
            x2, y1+radius,
            x2, y1+radius,
            x2, y2-radius,
            x2, y2-radius,
            x2, y2,
            x2-radius, y2,
            x2-radius, y2,
            x1+radius, y2,
            x1+radius, y2,
            x1, y2,
            x1, y2-radius,
            x1, y2-radius,
            x1, y1+radius,
            x1, y1+radius,
            x1, y1
        ]
        return self.create_polygon(points, **kwargs, smooth=True)

    def on_enter(self, e):
        self.itemconfig(self.rect, fill=self.hover_color)

    def on_leave(self, e):
        self.itemconfig(self.rect, fill=self.bg)

    def on_press(self, e):
        self.itemconfig(self.rect, fill="#555555")  # Darker when pressed

    def on_release(self, e):
        self.itemconfig(self.rect, fill=self.hover_color)
        if self.command:
            self.command()

class ProgressDialog(tk.Toplevel):
    def __init__(self, parent, title="處理中"):
        super().__init__(parent)
        self.title(title)
        self.configure(bg="#2a2a2a")
        self.resizable(False, False)
        self.transient(parent)
        self.grab_set()
        
        # Calculate position to center on parent
        parent_x = parent.winfo_rootx()
        parent_y = parent.winfo_rooty()
        parent_width = parent.winfo_width()
        parent_height = parent.winfo_height()
        
        width = 300
        height = 80
        x = parent_x + (parent_width - width) // 2
        y = parent_y + (parent_height - height) // 2
        
        self.geometry(f"{width}x{height}+{x}+{y}")
        
        self.label = tk.Label(self, text="正在處理檔案...", bg="#2a2a2a", fg="#ffffff", font=("SF Pro Text", 12))
        self.label.pack(pady=(10, 0))
        
        self.progress = ttk.Progressbar(self, orient="horizontal", mode="indeterminate", length=250)
        self.progress.pack(pady=15, padx=25)
        self.progress.start(10)
        
    def update_status(self, text):
        self.label.config(text=text)
        self.update()

class MacOSAlert(tk.Toplevel):
    def __init__(self, parent, title, message, icon_type="info"):
        super().__init__(parent)
        self.title("")
        self.configure(bg="#2a2a2a")
        self.resizable(False, False)
        self.transient(parent)
        
        # No window decorations on macOS-style dialogs
        self.overrideredirect(True)
        
        # Calculate position to center on parent
        parent_x = parent.winfo_rootx()
        parent_y = parent.winfo_rooty()
        parent_width = parent.winfo_width()
        parent_height = parent.winfo_height()
        
        width = 400
        height = 170
        x = parent_x + (parent_width - width) // 2
        y = parent_y + (parent_height - height) // 2
        
        self.geometry(f"{width}x{height}+{x}+{y}")
        
        # Add drop shadow effect (simulated with a frame)
        shadow_frame = tk.Frame(self, bg="#1a1a1a", bd=0)
        shadow_frame.place(x=3, y=3, width=width, height=height)
        
        # Main content frame
        main_frame = tk.Frame(self, bg="#2a2a2a", bd=0)
        main_frame.place(x=0, y=0, width=width, height=height)
        
        # Icon and title frame
        header_frame = tk.Frame(main_frame, bg="#2a2a2a", height=40)
        header_frame.pack(fill=tk.X, pady=(15, 5))
        
        # Icon (represented as text for simplicity)
        if icon_type == "info":
            icon_text = "ℹ️"
        elif icon_type == "warning":
            icon_text = "⚠️"
        elif icon_type == "error":
            icon_text = "❌"
        else:
            icon_text = "ℹ️"
            
        icon_label = tk.Label(header_frame, text=icon_text, bg="#2a2a2a", fg="#ffffff", font=("SF Pro Text", 24))
        icon_label.pack(side=tk.LEFT, padx=(20, 0))
        
        # Title
        title_label = tk.Label(header_frame, text=title, bg="#2a2a2a", fg="#ffffff", font=("SF Pro Text", 16, "bold"))
        title_label.pack(side=tk.LEFT, padx=10)
        
        # Message
        message_frame = tk.Frame(main_frame, bg="#2a2a2a")
        message_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=5)
        
        message_label = tk.Label(message_frame, text=message, bg="#2a2a2a", fg="#e0e0e0", 
                                font=("SF Pro Text", 12), wraplength=360, justify="left")
        message_label.pack(fill=tk.BOTH, expand=True)
        
        # Button frame
        button_frame = tk.Frame(main_frame, bg="#2a2a2a", height=50)
        button_frame.pack(fill=tk.X, pady=(0, 15))
        
        self.ok_button = MacOSButton(button_frame, text="確定", command=self.destroy, 
                                    width=80, height=30, bg="#0066cc", hover_color="#0077ee")
        self.ok_button.pack(side=tk.RIGHT, padx=(0, 20))
        
        # Make dialog modal
        self.grab_set()
        self.focus_set()
        
        # Position the window in the center of the parent
        self.update_idletasks()
        
        # Add a subtle bounce animation
        self.animate_entrance()
        
        # Bind Escape key to close
        self.bind("<Escape>", lambda e: self.destroy())
        
    def animate_entrance(self):
        # Store original position
        orig_y = self.winfo_y()
        
        # Start slightly above
        self.geometry(f"+{self.winfo_x()}+{orig_y-15}")
        self.update_idletasks()
        
        # Bounce down
        def bounce_down():
            for i in range(15):
                self.geometry(f"+{self.winfo_x()}+{self.winfo_y()+1}")
                self.update_idletasks()
                time.sleep(0.01)
        
        threading.Thread(target=bounce_down, daemon=True).start()

def extract_data(filepath):
    with open(filepath, 'r', encoding='utf-8') as f:
        lines = f.readlines()

    data = []
    i = 0
    while i < len(lines):
        line = lines[i].strip()

        # 偵測 Item 標題列
        match = re.match(r"#+\s*(.+?)\s*#+$", line)
        if match:
            base_item = match.group(1).strip()

            # 找下一個 Item 的邊界
            next_item_index = i + 1
            while next_item_index < len(lines):
                if re.match(r"#+\s*(.+?)\s*#+$", lines[next_item_index]):
                    break
                next_item_index += 1

            # 尋找參數格式或特殊格式
            found = False
            for j in range(i + 1, next_item_index):
                detail_line = lines[j].strip().rstrip(".")

                # --- 原本條件判斷 (保留不動) ---
                if (
                    "(" in detail_line and ")" in detail_line
                    and "(S)" not in detail_line
                    and "[" not in detail_line
                    and "]" not in detail_line
                    and re.search(r"\([^\)]*?\)\s*$", detail_line)
                ):
                    row_data = {h: "" for h in HEADERS}
                    row_data["Item"] = base_item
                    row_data["Detail"] = re.sub(r"\([^\)]*\)\s*$", "", detail_line).strip()
                    param_match = re.search(r"\(([^\)]*?)\)\s*$", detail_line)
                    if param_match:
                        all_kv_part = param_match.group(1)
                        t_match = re.search(r"T=\s*([\d\.]+mS)", all_kv_part, re.IGNORECASE)
                        if t_match:
                            row_data["T"] = t_match.group(1)
                        for k in keyword_to_header:
                            kv_match = re.search(rf"{re.escape(k)}=([^;,\s\)]+)", all_kv_part)
                            if kv_match:
                                row_data[k] = kv_match.group(1)
                    data.append(row_data)
                    found = True
                    break

                # --- ✅ 新增判斷：符合結構 xxx[...] Measure Check(yyy); SPEC=... ---
                elif re.search(r"\[.*?\].*?Measure Check\([^\)]+\)", detail_line):
                    detail_part = re.sub(r"\(.*?\)\s*;.*", "", detail_line).strip()
                    value_match = re.search(r"Measure Check\(([^)]+)\)", detail_line)
                    if value_match:
                        val = value_match.group(1).strip()
                        row_data = {h: "" for h in HEADERS}
                        row_data["Item"] = base_item
                        row_data["Detail"] = detail_part
                        row_data["I"] = val
                        data.append(row_data)
                        found = True  # 記錄已處理但不跳出，允許多筆連續處理
            # 若沒找到任何 Detail/參數，仍要記錄該 Item
            if not found:
                row_data = {h: "" for h in HEADERS}
                row_data["Item"] = base_item
                data.append(row_data)

            i = next_item_index
        else:
            i += 1

    return data

def save_to_excel(data, save_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Extracted Data"

    # 設定標題列
    for col, header in enumerate(HEADERS, start=1):
        ws[f"{get_column_letter(col)}1"] = header

    # 寫入資料
    for row_idx, row_data in enumerate(data, start=2):
        for col_idx, header in enumerate(HEADERS, start=1):
            ws[f"{get_column_letter(col_idx)}{row_idx}"] = row_data.get(header, "")

    wb.save(save_path)

class MainApplication(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("CP Rawdata 轉換工具")
        self.configure(bg="#1e1e1e")  # Dark background
        self.minsize(550, 300)
        self.geometry("550x300")
        
        # Set icon (would be replaced with actual file in production)
        # self.iconbitmap("icon.ico")  # For Windows
        # self.iconphoto(True, tk.PhotoImage(file="icon.png"))  # For Linux/macOS
        
        # Set system font
        self.system_font = font.nametofont("TkDefaultFont")
        self.system_font.configure(family="SF Pro Text", size=12)
        
        # Setting ttk style
        self.style = ttk.Style()
        self.style.theme_use('default')
        self.style.configure("TProgressbar", thickness=6, background='#0066cc')
        
        self.create_widgets()
        self.center_window()
        self.current_file = None
        
    def create_widgets(self):
        # Main frame with padding
        main_frame = tk.Frame(self, bg="#1e1e1e", padx=20, pady=20)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Header Frame
        header_frame = tk.Frame(main_frame, bg="#1e1e1e")
        header_frame.pack(fill=tk.X, pady=(0, 15))
        
        # Title with SF Pro Display font
        title_label = tk.Label(header_frame, text="CP Rawdata 擷取器", 
                               font=("SF Pro Display", 18, "bold"), 
                               bg="#1e1e1e", fg="#ffffff")
        title_label.pack(side=tk.LEFT)
        
        # Information icon with tooltip
        info_frame = tk.Frame(header_frame, bg="#1e1e1e")
        info_frame.pack(side=tk.RIGHT, padx=5)
        
        info_label = tk.Label(info_frame, text="ⓘ", font=("SF Pro Text", 14), 
                             bg="#1e1e1e", fg="#4a90e2", cursor="hand2")
        info_label.pack()
        
        # Tooltip for info button
        TooltipBase(info_label, "格式範例：\n############ Item文字 ############\nDetail名稱(參數1=值1, 參數2=值2...)")
        
        # File selection frame
        file_frame = tk.Frame(main_frame, bg="#1e1e1e")
        file_frame.pack(fill=tk.X, pady=10)
        
        # File path display with ellipsis for long paths
        self.file_var = tk.StringVar()
        self.file_var.set("尚未選擇檔案")
        
        file_label = tk.Label(file_frame, text="檔案路徑:", bg="#1e1e1e", fg="#e0e0e0", 
                             font=("SF Pro Text", 12))
        file_label.pack(side=tk.LEFT, padx=(0, 5))
        
        self.path_label = tk.Label(file_frame, textvariable=self.file_var, bg="#1e1e1e", 
                                  fg="#cccccc", anchor="w", width=40, font=("SF Pro Text", 12))
        self.path_label.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        # Button frame
        button_frame = tk.Frame(main_frame, bg="#1e1e1e")
        button_frame.pack(fill=tk.X, pady=20)
        
        # Open file button
        self.open_button = MacOSButton(button_frame, text="選擇 TXT 檔案", command=self.select_file,
                                      width=150, height=34, bg="#333333", hover_color="#404040")
        self.open_button.pack(side=tk.LEFT, padx=(0, 10))
        
        # Export button
        self.export_button = MacOSButton(button_frame, text="匯出 Excel", command=self.export_excel,
                                        width=150, height=34, bg="#0066cc", hover_color="#0077ee")
        self.export_button.pack(side=tk.LEFT)
        self.export_button.configure(state=tk.DISABLED)  # Initially disabled
        
        # Status frame at the bottom
        status_frame = tk.Frame(main_frame, bg="#252525", bd=0, height=30)
        status_frame.pack(fill=tk.X, side=tk.BOTTOM, pady=(20, 0))
        
        self.status_var = tk.StringVar()
        self.status_var.set("準備就緒")
        
        status_label = tk.Label(status_frame, textvariable=self.status_var, bg="#252525", 
                               fg="#a0a0a0", anchor="w", padx=10, pady=5, font=("SF Pro Text", 10))
        status_label.pack(fill=tk.X)

    def center_window(self):
        """Center the window on the screen"""
        self.update_idletasks()
        width = self.winfo_width()
        height = self.winfo_height()
        x = (self.winfo_screenwidth() // 2) - (width // 2)
        y = (self.winfo_screenheight() // 2) - (height // 2)
        self.geometry(f'{width}x{height}+{x}+{y}')
    
    def select_file(self):
        filepath = filedialog.askopenfilename(
            title="選擇 TXT 檔案",
            filetypes=[("Text files", "*.txt"), ("All files", "*.*")],
            initialdir=os.path.expanduser("~/Documents")  # Default to Documents folder
        )
        
        if not filepath:
            return
            
        # Update file path display with ellipsis for long paths
        if len(filepath) > 40:
            display_path = filepath[:18] + "..." + filepath[-19:]
        else:
            display_path = filepath
            
        self.file_var.set(display_path)
        self.current_file = filepath
        
        # Enable export button
        self.export_button.configure(state=tk.NORMAL)
        self.status_var.set(f"已載入: {Path(filepath).name}")

    def process_file(self, filepath, save_path, progress_dialog):
        try:
            # Update status
            progress_dialog.update_status("正在解析檔案...")
            
            # Extract data
            data = extract_data(filepath)
            
            if not data:
                # Close progress dialog before showing error
                progress_dialog.destroy()
                MacOSAlert(self, "無結果", "檔案中未找到符合格式的文字。", "warning")
                return
            
            # Update status
            progress_dialog.update_status("正在儲存到 Excel...")
            
            # Save to Excel
            save_to_excel(data, save_path)
            
            # Close progress dialog
            progress_dialog.destroy()
            
            # Show success message
            success_message = f"已成功擷取 {len(data)} 筆資料並儲存至：\n{Path(save_path).name}"
            MacOSAlert(self, "完成", success_message, "info")
            
            # Update status
            self.status_var.set(f"已匯出 {len(data)} 筆資料")
            
        except Exception as e:
            # Close progress dialog before showing error
            progress_dialog.destroy()
            MacOSAlert(self, "錯誤", str(e), "error")
            self.status_var.set("處理時發生錯誤")

    def export_excel(self):
        if not self.current_file:
            MacOSAlert(self, "注意", "請先選擇一個 TXT 檔案。", "warning")
            return
            
        # Default filename based on input file
        default_filename = Path(self.current_file).stem + ".xlsx"
        
        save_path = filedialog.asksaveasfilename(
            title="儲存 Excel 檔案",
            defaultextension=".xlsx", 
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=default_filename,
            initialdir=Path(self.current_file).parent
        )
        
        if not save_path:
            return
            
        # Show progress dialog
        progress_dialog = ProgressDialog(self)
        
        # Process in a separate thread to keep UI responsive
        threading.Thread(
            target=self.process_file, 
            args=(self.current_file, save_path, progress_dialog),
            daemon=True
        ).start()

# Set macOS-style appearance for the application
def set_macos_appearance():
    try:
        # For macOS
        import platform
        if platform.system() == 'Darwin':
            os.system('''defaults write -g NSRequiresAquaSystemAppearance -bool YES''')
            from tkmacosx import ColorVar, ColorscaleVar
    except:
        pass

if __name__ == "__main__":
    set_macos_appearance()
    app = MainApplication()
    app.mainloop()